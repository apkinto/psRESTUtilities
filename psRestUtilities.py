import json
import requests
import datetime
import xml.etree.ElementTree as ET
import sys, os
import urllib.parse
import csv
import logging
from openpyxl import load_workbook


def getTime():
	currentTime = datetime.datetime.now()
	return currentTime

def setVariables( config ):
	'''
		Sets variables from psPythonConfig.xml.   Currently assumes it is in the PS bin directory.
	'''
	variable = {}
	config = ET.parse(config)
	root = config.getroot()
	for var in root.find('variableList'):
		variable[var.tag] = var.text
	return variable

def setLogging():
	logger = logging.getLogger(__name__)
	logger.setLevel(logging.DEBUG)
	
	fh = logging.FileHandler('psPython.log')
	fh.setLevel(logging.INFO)
	
	ch = logging.StreamHandler()
	ch.setLevel(logging.INFO)
	
	formatter = logging.Formatter('%(asctime)s %(name)s  %(levelname)s \t %(message)s')
	fh.setFormatter(formatter)
	ch.setFormatter(formatter)
	
	logger.addHandler(fh)
	logger.addHandler(ch)
	return logger
		
def getRest( url, session, payload, requestHeader, authorization, recordLimit, log ):
	#log = setLogging()
	payload = ''
	querystring = { 
					"limit": recordLimit 
					}
	
	start = getTime()
	try:
		r = session.get( url, data=payload, headers=requestHeader, params=querystring, auth=authorization )
		data = r.content
		output = json.loads(data)
		end = getTime()
		time = end - start
		#log.info('\t\tStatusCode: %s\t%s sec\t%s' % (r.status_code, time, url))
	except:
		output = {'items' : None}
		r.status_code
		end = getTime()
		time = end - start
		log.info('   **ERROR**')
	
	return output, time, r.status_code

def postRest( url, session, body, requestHeader, authorization, log ):
	#log = setLogging()
	start = getTime()
	try:
		r = session.post( url, json=body, headers=requestHeader, auth=authorization )
		print ( r.status_code, r.text )
		end = getTime()
		time = end - start
	except:
		r.status_code
		end = getTime()
		time = end - start
		log.info('   **ERROR**')
	
	return time, r.status_code, r.text

		
def getResources( filename ):
	resourceNames = []
	with open(filename, 'r', newline = '') as f:  
		for line in f:	
			resourceNames.append(line.rstrip())
	return resourceNames

def getPsPlanId ( psPlanOutput, log ):
	psPlans = []
	for p in psPlanOutput['items']:
		psPlans.append( { 
							'PlanId' : p['PlanId'],
							'PlaName': p['PlanName'] 
						} 
						)
	log.info('--> Fetching Data for following Plans (PlanID): %s\n' % ( psPlans ) )
	return psPlans
	
def idCode( output, entity, log ):
	objectIdCode = {}
	entityId = entity + 'Id'
	entityCode = entity + 'Code'
	
	for o in output['items']:
		objectIdCode[o[entityCode]] = o[entityId]
	log.info('\t\tCode to Id mapping for\t %s : %s mapping\n' % ( entityCode, entityId ) )
	return objectIdCode
	
	
def writeCsv ( list, filename, outDir ):
	file = filename + '.csv'
	csvFile = os.path.join( outDir, file)
	with open( csvFile, 'w', newline = '' ) as f:
		header = []
		for h in list[0].keys():
			header.append( h )
		csvwriter= csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
		csvwriter.writerow( header )
		for i in list:
			w = csv.DictWriter(f, i.keys())
			w.writerow(i)
		f.close()

def getUrl ( *n ):
	newUrl = '/'.join( n )
	
	return newUrl
	
def getJsonItems ( jsonOutput ):
	objectList = jsonOutput['items']
	
	return objectList

def scmAuth ( user, password ):
	r = requests.Session()
	r.auth = ( user, password )
	#r.headers = {'Content-type': 'application/json', 'REST-Framework-Version': '1'}
	r.headers={	'Cache-Control': 'no-cache','Content-Type': 'application/json', 'REST-Framework-Version': '1', 'Connection': 'close', 	 }
	payload = ''
	
	return r, r.auth, r.headers, payload
	
def readExcel ( filename, object ) :
	wb = load_workbook( filename )
	sheet = wb[ object ]
	res = list( sheet )  				# list of records in excel sheet
	final = []							# List of records as dictionary
	
	for x in range(1, sheet.max_row ):
		partFinal = {}	
		for y in range (0, sheet.max_column):
			'''  If date, Change from datetime to ISO date format '''
			if isinstance(res[x][y].value, datetime.datetime):
				partFinal[res[0][y].value] = res[x][y].value.replace(tzinfo=datetime.timezone.utc).isoformat()
			else:
				partFinal[res[0][y].value] = res[x][y].value
		final.append(partFinal)
	
	return final
	